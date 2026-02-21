#!/usr/bin/env python3
"""
根据 ap.json 中的 MAC 地址匹配数据，为 Excel 文件追加 4 列：
  - 网关name
  - 上线方式 (ap.uplink)
  - app版本 (container.apps[0].name + container.apps[0].version)
  - group
"""

import json
import openpyxl

# ---- 配置 ----
AP_JSON_PATH = "ap.json"
EXCEL_INPUT = "UE EMMC report_summary 20260212 1512.xlsx"
EXCEL_OUTPUT = "UE EMMC report_summary 20260212 1512_enriched.xlsx"
MAC_COLUMN = 2  # Column B

# ---- 1. 加载 ap.json，构建 MAC 查找字典 ----
with open(AP_JSON_PATH, "r", encoding="utf-8") as f:
    ap_list = json.load(f)

ap_dict: dict[str, dict] = {}
for item in ap_list:
    mac = item.get("mac", "").upper()
    if mac:
        ap_dict[mac] = item

print(f"ap.json 加载完成，共 {len(ap_dict)} 条记录")

# ---- 2. 加载 Excel ----
wb = openpyxl.load_workbook(EXCEL_INPUT)
ws = wb.active
max_col = ws.max_column

# ---- 3. 写入新表头 ----
header_col_start = max_col + 1
ws.cell(row=1, column=header_col_start).value = "网关name"
ws.cell(row=1, column=header_col_start + 1).value = "上线方式"
ws.cell(row=1, column=header_col_start + 2).value = "app版本"
ws.cell(row=1, column=header_col_start + 3).value = "group"

# ---- 4. 遍历数据行，匹配并填充 ----
matched = 0
not_found = 0

for row in range(2, ws.max_row + 1):
    mac_val = ws.cell(row=row, column=MAC_COLUMN).value
    if not mac_val:
        continue

    mac_key = str(mac_val).strip().upper()
    ap_info = ap_dict.get(mac_key)

    if ap_info is None:
        not_found += 1
        continue

    matched += 1

    # 网关name
    ws.cell(row=row, column=header_col_start).value = ap_info.get("name", "")

    # 上线方式: ap.uplink
    ap_field = ap_info.get("ap", {})
    uplink = ap_field.get("uplink", "") if isinstance(ap_field, dict) else ""
    ws.cell(row=row, column=header_col_start + 1).value = uplink

    # app版本: container.apps[0].name + container.apps[0].version
    container = ap_info.get("container", {})
    apps = container.get("apps", []) if isinstance(container, dict) else []
    if apps and len(apps) > 0:
        app_name = apps[0].get("name", "")
        app_version = apps[0].get("version", "")
        ws.cell(row=row, column=header_col_start + 2).value = f"{app_name}{app_version}"
    else:
        ws.cell(row=row, column=header_col_start + 2).value = ""

    # group
    ws.cell(row=row, column=header_col_start + 3).value = ap_info.get("group", "")

# ---- 5. 保存 ----
wb.save(EXCEL_OUTPUT)
print(f"处理完成！匹配成功: {matched}, 未找到: {not_found}")
print(f"输出文件: {EXCEL_OUTPUT}")
